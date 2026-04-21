from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from scrapekit.exporters.base import BaseExporter


class ExcelExporter(BaseExporter):
    """Excel exporter with auto-column-width and styled headers."""

    extension = "xlsx"

    def __init__(
        self,
        output_dir: str | Path = "./output",
        sheet_name: str = "Data",
    ) -> None:
        super().__init__(output_dir)
        self._sheet_name = sheet_name

    def _write(self, df: pd.DataFrame, path: Path) -> None:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=self._sheet_name)
            self._style_worksheet(writer.sheets[self._sheet_name], df)

    def _style_worksheet(self, ws: object, df: pd.DataFrame) -> None:
        from openpyxl.worksheet.worksheet import Worksheet
        ws: Worksheet  # type: ignore[no-redef]

        header_fill = PatternFill("solid", fgColor="2563EB")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, _ in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for col_idx, column in enumerate(df.columns, start=1):
            max_len = max(
                df[column].astype(str).map(len).max(),
                len(str(column)),
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)
